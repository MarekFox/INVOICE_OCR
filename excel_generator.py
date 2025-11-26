"""
FAKTURA BOT v5.0 - Excel Report Generator
==========================================
Profesjonalny generator raportów Excel z wykresami i formatowaniem
"""

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side,
    NamedStyle, Color, Protection
)
from openpyxl.chart import PieChart, BarChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule
from typing import List, Dict, Any, Optional
from datetime import datetime
from decimal import Decimal
import logging

from parsers import ParsedInvoice
from config import CONFIG

logger = logging.getLogger(__name__)

class ExcelReportGenerator:
    """Generator profesjonalnych raportów Excel"""
    
    # Style kolorów
    COLORS = {
        'header_blue': 'FF0070C0',
        'header_green': 'FF70AD47',
        'header_orange': 'FFFFC000',
        'light_blue': 'FFDBEEF3',
        'light_green': 'FFE2EFDA',
        'light_yellow': 'FFFFF2CC',
        'error_red': 'FFFFCCCC',
        'warning_yellow': 'FFFFFFCC',
        'success_green': 'FFCCFFCC'
    }
    
    def __init__(self, filename: str = None):
        self.filename = filename or f"Raport_Faktur_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        self.wb = Workbook()
        self.setup_styles()
        
    def setup_styles(self):
        """Konfiguruje style dla dokumentu"""
        # Styl nagłówka
        header_style = NamedStyle(name="header_style")
        header_style.font = Font(bold=True, color="FFFFFFFF", size=11)
        header_style.fill = PatternFill(start_color=self.COLORS['header_blue'],
                                       end_color=self.COLORS['header_blue'],
                                       fill_type="solid")
        header_style.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        header_style.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='medium')
        )
        self.wb.add_named_style(header_style)
        
        # Styl sum
        total_style = NamedStyle(name="total_style")
        total_style.font = Font(bold=True, size=11)
        total_style.fill = PatternFill(start_color=self.COLORS['light_blue'],
                                      end_color=self.COLORS['light_blue'],
                                      fill_type="solid")
        total_style.border = Border(
            top=Side(style='double'),
            bottom=Side(style='double')
        )
        self.wb.add_named_style(total_style)
        
        # Styl waluty
        currency_style = NamedStyle(name="currency_style")
        currency_style.number_format = '#,##0.00 zł'
        currency_style.alignment = Alignment(horizontal="right")
        self.wb.add_named_style(currency_style)
        
    def generate(self, invoices: List[ParsedInvoice], options: Dict = None):
        """
        Generuje kompletny raport
        
        Args:
            invoices: Lista sparsowanych faktur
            options: Opcje generowania raportu
        """
        if not invoices:
            logger.warning("Brak faktur do wygenerowania raportu")
            return
            
        options = options or {}
        
        # Usuń domyślny arkusz
        if "Sheet" in self.wb.sheetnames:
            self.wb.remove(self.wb["Sheet"])
            
        # Generuj arkusze
        self._create_summary_sheet(invoices)
        self._create_details_sheet(invoices)
        self._create_items_sheet(invoices)
        self._create_statistics_sheet(invoices)
        
        if options.get('include_charts', True):
            self._create_charts_sheet(invoices)
            
        if options.get('include_pivot', False):
            self._create_pivot_sheet(invoices)
            
        if options.get('include_validation', True):
            self._create_validation_sheet(invoices)
            
        # Ustaw arkusz podsumowania jako aktywny
        self.wb.active = self.wb['Podsumowanie']
        
        # Zapisz plik
        self.save()
        
    def _create_summary_sheet(self, invoices: List[ParsedInvoice]):
        """Tworzy arkusz podsumowania"""
        from utils import DateUtils
        ws = self.wb.create_sheet("Podsumowanie")
        
        # Nagłówek
        ws.append([
            "LP", "Nr Faktury", "Typ", 
            "Data Wyst.", "Data Sprz.", "Termin Płatn.",  # ← DODANE: Data Sprz.
            "Dostawca", "NIP Dostawcy", 
            "Nabywca", "NIP Nabywcy",
            "Konto", 
            "Netto", "VAT", "Brutto", "Waluta", 
            "Status", "Uwagi"
        ])
        
        # Formatuj nagłówek
        for cell in ws[1]:
            cell.style = "header_style"
            
        # Dane
        for i, invoice in enumerate(invoices, 1):
            status = self._get_invoice_status(invoice)
            warnings = ', '.join(invoice.parsing_warnings[:3])
            
            bank_account = invoice.supplier_accounts[0] if invoice.supplier_accounts else "Brak"
            
            row = [
                i,
                invoice.invoice_id,
                invoice.invoice_type,
                DateUtils.format_date_output(invoice.issue_date),   # ← ZMIENIONE: dd.mm.rrrr
                DateUtils.format_date_output(invoice.sale_date),    # ← ZMIENIONE: dd.mm.rrrr
                DateUtils.format_date_output(invoice.due_date),     # ← ZMIENIONE: dd.mm.rrrr
                invoice.supplier_name[:50],
                invoice.supplier_tax_id,
                invoice.buyer_name[:50],
                invoice.buyer_tax_id,
                bank_account,
                float(invoice.total_net),
                float(invoice.total_vat),
                float(invoice.total_gross),
                invoice.currency,
                status,
                warnings
            ]
            ws.append(row)
            
            # Formatowanie warunkowe dla statusu
            row_num = ws.max_row
            status_cell = ws.cell(row=row_num, column=13)
            
            if status == "✅ OK":
                status_cell.fill = PatternFill(start_color=self.COLORS['success_green'],
                                              end_color=self.COLORS['success_green'],
                                              fill_type="solid")
            elif "⚠️" in status:
                status_cell.fill = PatternFill(start_color=self.COLORS['warning_yellow'],
                                              end_color=self.COLORS['warning_yellow'],
                                              fill_type="solid")
            elif "❌" in status:
                status_cell.fill = PatternFill(start_color=self.COLORS['error_red'],
                                              end_color=self.COLORS['error_red'],
                                              fill_type="solid")
        
        # Wiersz sum
        ws.append([])  # Pusty wiersz
        total_row = ws.max_row + 1
        ws.append([
            "", "", "", "", "", "", "", "SUMA:",
            f"=SUM(I2:I{ws.max_row-2})",
            f"=SUM(J2:J{ws.max_row-2})",
            f"=SUM(K2:K{ws.max_row-2})",
            "", "", ""
        ])
        
        # Formatuj wiersz sum
        for col in range(8, 12):
            cell = ws.cell(row=total_row, column=col)
            cell.style = "total_style"
            
        # Formatuj kolumny z kwotami
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row-1, min_col=9, max_col=11):
            for cell in row:
                cell.number_format = '#,##0.00 zł'
                
        # Ustaw szerokości kolumn
        column_widths = {
            'A': 5,   # LP
            'B': 20,  # Nr Faktury
            'C': 12,  # Typ
            'D': 12,  # Data
            'E': 30,  # Dostawca
            'F': 15,  # NIP
            'G': 30,  # Nabywca
            'H': 15,  # NIP
            'I': 15,  # Netto
            'J': 12,  # VAT
            'K': 15,  # Brutto
            'L': 8,   # Waluta
            'M': 12,  # Status
            'N': 40   # Uwagi
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
            
        # Dodaj filtry
        ws.auto_filter.ref = f"A1:N{ws.max_row-2}"
        
        # Zablokuj nagłówek
        ws.freeze_panes = 'A2'
        
    def _create_details_sheet(self, invoices: List[ParsedInvoice]):
        """Tworzy arkusz ze szczegółami"""
        ws = self.wb.create_sheet("Szczegóły")
        
        for i, invoice in enumerate(invoices):
            if i > 0:
                ws.append([])  # Separator między fakturami
                
            # Nagłówek faktury
            ws.append([f"FAKTURA: {invoice.invoice_id}"])
            ws[ws.max_row][0].font = Font(bold=True, size=14)
            ws[ws.max_row][0].fill = PatternFill(start_color=self.COLORS['header_green'],
                                                end_color=self.COLORS['header_green'],
                                                fill_type="solid")
            
            # Dane faktury
            details = [
                ["Typ dokumentu:", invoice.invoice_type],
                ["Data wystawienia:", invoice.issue_date.strftime('%Y-%m-%d')],
                ["Data sprzedaży:", invoice.sale_date.strftime('%Y-%m-%d')],
                ["Termin płatności:", invoice.due_date.strftime('%Y-%m-%d')],
                ["", ""],
                ["SPRZEDAWCA", ""],
                ["Nazwa:", invoice.supplier_name],
                ["NIP:", invoice.supplier_tax_id],
                ["Adres:", invoice.supplier_address],
                ["Konto:", invoice.supplier_accounts[0] if invoice.supplier_accounts else "Brak"],
                ["", ""],
                ["NABYWCA", ""],
                ["Nazwa:", invoice.buyer_name],
                ["NIP:", invoice.buyer_tax_id],
                ["Adres:", invoice.buyer_address],
                ["", ""],
                ["PODSUMOWANIE", ""],
                ["Wartość netto:", f"{invoice.total_net:.2f} {invoice.currency}"],
                ["VAT:", f"{invoice.total_vat:.2f} {invoice.currency}"],
                ["Wartość brutto:", f"{invoice.total_gross:.2f} {invoice.currency}"],
                ["", ""],
                ["Metoda płatności:", invoice.payment_method],
                ["Status płatności:", invoice.payment_status]
            ]
            
            for row in details:
                ws.append(row)
                # Formatowanie nagłówków sekcji
                if row[0] in ["SPRZEDAWCA", "NABYWCA", "PODSUMOWANIE"]:
                    ws[ws.max_row][0].font = Font(bold=True)
                    ws[ws.max_row][0].fill = PatternFill(
                        start_color=self.COLORS['light_blue'],
                        end_color=self.COLORS['light_blue'],
                        fill_type="solid"
                    )
                    
            # Pozycje faktury
            if invoice.line_items:
                ws.append([])
                ws.append(["POZYCJE FAKTURY"])
                ws[ws.max_row][0].font = Font(bold=True)
                
                ws.append(["LP", "Opis", "Ilość", "Cena jedn.", "Wartość"])
                for cell in ws[ws.max_row]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color=self.COLORS['light_green'],
                                           end_color=self.COLORS['light_green'],
                                           fill_type="solid")
                
                for j, item in enumerate(invoice.line_items, 1):
                    ws.append([
                        j,
                        item.get('description', ''),
                        item.get('quantity', 0),
                        item.get('unit_price', 0),
                        item.get('total', 0)
                    ])
                    
        # Ustaw szerokości kolumn
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 50
        
    def _create_items_sheet(self, invoices: List[ParsedInvoice]):
        """Tworzy arkusz ze wszystkimi pozycjami"""
        ws = self.wb.create_sheet("Pozycje")
        
        # Nagłówek
        ws.append([
            "Nr Faktury", "Data", "Dostawca", "LP", "Opis", 
            "Ilość", "Cena jedn.", "Wartość netto", "VAT", "Wartość brutto"
        ])
        
        for cell in ws[1]:
            cell.style = "header_style"
            
        # Dane
        for invoice in invoices:
            for i, item in enumerate(invoice.line_items, 1):
                # Oblicz VAT (zakładamy 23% jeśli nie podano)
                total = Decimal(str(item.get('total', 0)))
                vat_amount = total - (total / Decimal('1.23'))
                net_amount = total - vat_amount
                
                ws.append([
                    invoice.invoice_id,
                    invoice.issue_date.strftime('%Y-%m-%d'),
                    invoice.supplier_name[:30],
                    i,
                    item.get('description', '')[:100],
                    item.get('quantity', 0),
                    float(item.get('unit_price', 0)),
                    float(net_amount),
                    float(vat_amount),
                    float(total)
                ])
                
        # Formatuj kolumny liczbowe
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=6, max_col=10):
            row[0].number_format = '0'  # Ilość
            for cell in row[1:]:
                cell.number_format = '#,##0.00 zł'
                
        # Ustaw szerokości kolumn
        column_widths = {
            'A': 20, 'B': 12, 'C': 25, 'D': 5, 'E': 50,
            'F': 10, 'G': 15, 'H': 15, 'I': 12, 'J': 15
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
            
        # Dodaj tabelę
        if ws.max_row > 1:
            table = Table(displayName="TabelaPozycji", ref=f"A1:J{ws.max_row}")
            style = TableStyleInfo(
                name="TableStyleMedium2", showFirstColumn=False,
                showLastColumn=False, showRowStripes=True, showColumnStripes=False
            )
            table.tableStyleInfo = style
            ws.add_table(table)
            
    def _create_statistics_sheet(self, invoices: List[ParsedInvoice]):
        """Tworzy arkusz ze statystykami"""
        ws = self.wb.create_sheet("Statystyki")
        
        # Oblicz statystyki
        stats = self._calculate_statistics(invoices)
        
        # Wyświetl statystyki
        ws.append(["STATYSTYKI OGÓLNE"])
        ws[1][0].font = Font(bold=True, size=14)
        ws.append([])
        
        general_stats = [
            ["Liczba faktur:", stats['total_count']],
            ["Łączna wartość netto:", f"{stats['total_net']:.2f} PLN"],
            ["Łączny VAT:", f"{stats['total_vat']:.2f} PLN"],
            ["Łączna wartość brutto:", f"{stats['total_gross']:.2f} PLN"],
            ["", ""],
            ["Średnia wartość faktury:", f"{stats['avg_invoice_value']:.2f} PLN"],
            ["Mediana wartości:", f"{stats['median_invoice_value']:.2f} PLN"],
            ["Najwyższa faktura:", f"{stats['max_invoice_value']:.2f} PLN"],
            ["Najniższa faktura:", f"{stats['min_invoice_value']:.2f} PLN"],
            ["", ""],
            ["Liczba dostawców:", stats['unique_suppliers']],
            ["Liczba nabywców:", stats['unique_buyers']],
            ["", ""],
            ["Faktury OK:", stats['valid_invoices']],
            ["Faktury z błędami:", stats['error_invoices']],
            ["Faktury z ostrzeżeniami:", stats['warning_invoices']]
        ]
        
        for row in general_stats:
            ws.append(row)
            
        # TOP Dostawcy
        ws.append([])
        ws.append(["TOP 10 DOSTAWCÓW"])
        ws[ws.max_row][0].font = Font(bold=True, size=12)
        ws.append(["Dostawca", "Liczba faktur", "Wartość brutto"])
        
        for supplier in stats['top_suppliers'][:10]:
            ws.append([supplier['name'], supplier['count'], f"{supplier['total']:.2f} PLN"])
            
        # Podsumowanie miesięczne
        ws.append([])
        ws.append(["PODSUMOWANIE MIESIĘCZNE"])
        ws[ws.max_row][0].font = Font(bold=True, size=12)
        ws.append(["Miesiąc", "Liczba faktur", "Wartość brutto"])
        
        for month in stats['monthly_summary']:
            ws.append([month['month'], month['count'], f"{month['total']:.2f} PLN"])
            
        # Formatowanie
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        
    def _create_charts_sheet(self, invoices: List[ParsedInvoice]):
        """Tworzy arkusz z wykresami"""
        ws = self.wb.create_sheet("Wykresy")
        
        # Przygotuj dane do wykresów
        stats = self._calculate_statistics(invoices)
        
        # Dane dla wykresu kołowego - podział na dostawców
        ws.append(["Dostawca", "Wartość"])
        for supplier in stats['top_suppliers'][:5]:
            ws.append([supplier['name'], supplier['total']])
        others_total = sum(s['total'] for s in stats['top_suppliers'][5:])
        if others_total > 0:
            ws.append(["Pozostali", others_total])
            
        # Wykres kołowy
        pie = PieChart()
        labels = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
        data = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        pie.title = "Top 5 Dostawców"
        
        # Dodaj etykiety danych
        pie.dataLabels = DataLabelList()
        pie.dataLabels.showPercent = True
        
        ws.add_chart(pie, "D2")
        
        # Dane dla wykresu słupkowego - miesięczne
        start_row = ws.max_row + 3
        ws.cell(row=start_row, column=1, value="Miesiąc")
        ws.cell(row=start_row, column=2, value="Wartość")
        
        for month in stats['monthly_summary']:
            ws.append([month['month'], month['total']])
            
        # Wykres słupkowy
        bar = BarChart()
        bar.type = "col"
        bar.style = 10
        bar.title = "Wartość faktur miesięcznie"
        bar.y_axis.title = "Wartość (PLN)"
        bar.x_axis.title = "Miesiąc"
        
        data = Reference(ws, min_col=2, min_row=start_row, max_row=ws.max_row)
        categories = Reference(ws, min_col=1, min_row=start_row+1, max_row=ws.max_row)
        bar.add_data(data, titles_from_data=True)
        bar.set_categories(categories)
        
        ws.add_chart(bar, "D20")
        
    def _create_validation_sheet(self, invoices: List[ParsedInvoice]):
        """Tworzy arkusz z wynikami walidacji"""
        ws = self.wb.create_sheet("Walidacja")
        
        # Nagłówek
        ws.append([
            "Nr Faktury", "Status", "Poziom pewności", "Błędy", "Ostrzeżenia", "Uwagi"
        ])
        
        for cell in ws[1]:
            cell.style = "header_style"
            
        # Dane
        for invoice in invoices:
            status = "✅ OK" if invoice.is_verified else "❌ Błąd"
            errors = '; '.join(invoice.parsing_errors) if invoice.parsing_errors else "Brak"
            warnings = '; '.join(invoice.parsing_warnings) if invoice.parsing_warnings else "Brak"
            
            row = [
                invoice.invoice_id,
                status,
                f"{invoice.confidence:.1%}",
                errors,
                warnings,
                "Do weryfikacji" if invoice.confidence < 0.8 else ""
            ]
            ws.append(row)
            
            # Kolorowanie według statusu
            row_num = ws.max_row
            if status == "✅ OK":
                fill_color = self.COLORS['success_green']
            else:
                fill_color = self.COLORS['error_red']
                
            ws.cell(row=row_num, column=2).fill = PatternFill(
                start_color=fill_color,
                end_color=fill_color,
                fill_type="solid"
            )
            
        # Formatowanie warunkowe dla pewności
        ws.conditional_formatting.add(
            f'C2:C{ws.max_row}',
            ColorScaleRule(
                start_type='min', start_color='FFFF0000',  # Czerwony
                mid_type='percentile', mid_value=50, mid_color='FFFFFF00',  # Żółty  
                end_type='max', end_color='FF00FF00'  # Zielony
            )
        )
        
        # Ustaw szerokości kolumn
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 50
        ws.column_dimensions['E'].width = 50
        ws.column_dimensions['F'].width = 30
        
    def _create_pivot_sheet(self, invoices: List[ParsedInvoice]):
        """Tworzy arkusz z tabelą przestawną (wymaga danych)"""
        ws = self.wb.create_sheet("Analiza")
        
        # Przygotuj dane do tabeli przestawnej
        data = []
        for invoice in invoices:
            data.append({
                'Data': invoice.issue_date.strftime('%Y-%m'),
                'Dostawca': invoice.supplier_name[:30],
                'Kategoria': invoice.invoice_type,
                'Wartość netto': float(invoice.total_net),
                'VAT': float(invoice.total_vat),
                'Wartość brutto': float(invoice.total_gross)
            })
            
        df = pd.DataFrame(data)
        
        # Utworz tabelę przestawną
        pivot = pd.pivot_table(
            df,
            values=['Wartość netto', 'VAT', 'Wartość brutto'],
            index=['Data'],
            columns=['Kategoria'],
            aggfunc='sum',
            fill_value=0
        )
        
        # Zapisz do arkusza
        for r in dataframe_to_rows(pivot, index=True, header=True):
            ws.append(r)
            
        # Formatowanie
        for cell in ws[1]:
            cell.font = Font(bold=True)
            
    def _get_invoice_status(self, invoice: ParsedInvoice) -> str:
        """Określa status faktury"""
        if invoice.is_duplicate:
            return "🔄 Duplikat"
        elif invoice.parsing_errors:
            return f"❌ {len(invoice.parsing_errors)} błędów"
        elif invoice.parsing_warnings:
            return f"⚠️ {len(invoice.parsing_warnings)} ostrzeżeń"
        elif invoice.confidence < 0.8:
            return "❓ Niska pewność"
        else:
            return "✅ OK"
            
    def _calculate_statistics(self, invoices: List[ParsedInvoice]) -> Dict:
        """Oblicza statystyki dla faktur"""
        from collections import Counter
        import statistics
        
        # Podstawowe statystyki
        total_net = sum(float(inv.total_net) for inv in invoices)
        total_vat = sum(float(inv.total_vat) for inv in invoices)
        total_gross = sum(float(inv.total_gross) for inv in invoices)
        
        gross_values = [float(inv.total_gross) for inv in invoices]
        
        # Dostawcy
        supplier_totals = Counter()
        supplier_counts = Counter()
        
        for inv in invoices:
            supplier_totals[inv.supplier_name] += float(inv.total_gross)
            supplier_counts[inv.supplier_name] += 1
            
        top_suppliers = [
            {
                'name': name,
                'count': supplier_counts[name],
                'total': total
            }
            for name, total in supplier_totals.most_common(20)
        ]
        
        # Miesięczne
        monthly = Counter()
        monthly_counts = Counter()
        
        for inv in invoices:
            month_key = inv.issue_date.strftime('%Y-%m')
            monthly[month_key] += float(inv.total_gross)
            monthly_counts[month_key] += 1
            
        monthly_summary = [
            {
                'month': month,
                'count': monthly_counts[month],
                'total': total
            }
            for month, total in sorted(monthly.items())
        ]
        
        # Statusy
        valid = sum(1 for inv in invoices if inv.is_verified)
        errors = sum(1 for inv in invoices if inv.parsing_errors)
        warnings = sum(1 for inv in invoices if inv.parsing_warnings and not inv.parsing_errors)
        
        return {
            'total_count': len(invoices),
            'total_net': total_net,
            'total_vat': total_vat,
            'total_gross': total_gross,
            'avg_invoice_value': statistics.mean(gross_values) if gross_values else 0,
            'median_invoice_value': statistics.median(gross_values) if gross_values else 0,
            'max_invoice_value': max(gross_values) if gross_values else 0,
            'min_invoice_value': min(gross_values) if gross_values else 0,
            'unique_suppliers': len(supplier_totals),
            'unique_buyers': len(set(inv.buyer_name for inv in invoices)),
            'valid_invoices': valid,
            'error_invoices': errors,
            'warning_invoices': warnings,
            'top_suppliers': top_suppliers,
            'monthly_summary': monthly_summary
        }
        
    def save(self, filename: str = None):
        """Zapisuje plik Excel"""
        save_path = filename or self.filename
        self.wb.save(save_path)
        logger.info(f"Raport zapisany: {save_path}")
        return save_path